import os
import time
import json
import shutil
import telebot
import openpyxl
from dotenv import load_dotenv
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, Document


load_dotenv()
BOT_TOKEN = os.getenv('BOT_TOKEN')
DATA_FILE = 'data.json'




def load_data(filename):
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def save_data(filename, data):
    """Сохраняет данные в JSON."""
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def validate_xlsx(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        return sheet['A1'].value == "track" and sheet['B1'].value == "date"
    except Exception:
        return False

def move_file(file_path: str, target_dir: str) -> bool:
    """
    Перемещает файл в указанную директорию.
    
    :param file_path: Путь к файлу, который нужно переместить.
    :param target_dir: Путь к целевой директории.
    :return: True, если файл успешно перемещён, иначе False.
    """
    if not os.path.isfile(file_path):
        print(f"Ошибка: Файл '{file_path}' не существует.")
        return False

    if not os.path.exists(target_dir):
        os.makedirs(target_dir)  # Создаём директорию, если её нет

    try:
        shutil.move(file_path, os.path.join(target_dir, os.path.basename(file_path)))
        return True
    except Exception as e:
        print(f"Ошибка при перемещении файла: {e}")
        return False

text_data = load_data(DATA_FILE)
filename_list = text_data.get("filename", [])

users_data = text_data.get("users", {})
ADMINS = users_data.get("admins", [])
MODERATORS = users_data.get("moderators", [])
user_language = users_data.get("user_language", {})
waiting_for_admin = users_data.get("waiting_for_admin", {})


bot = telebot.TeleBot(BOT_TOKEN)
 
    
def lan_checker(message):
    if message.text == "Русский":
        return "ru"
    elif message.text == "Тоҷикӣ":
        return "tj"
    else:
        return None
    

def extract_keywords_from_excel(file_path: str) -> dict:
    """
    Читает Excel-файл и создаёт словарь из значений столбца 'track' и соответствующих им значений из 'date'.

    :param file_path: Путь к Excel-файлу (.xlsx).
    :return: Словарь {track: date}.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Проверяем, есть ли нужные заголовки
        if sheet["A1"].value != "track" or sheet["B1"].value != "date":
            raise ValueError("Ошибка! Ожидаемые заголовки: 'track' в колонке A, 'date' в колонке B.")

        data_dict = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Читаем со 2-й строки (пропуская заголовки)
            track, date = row
            if track:  # Игнорируем пустые строки
                data_dict[str(track)] = date
                print(data_dict[str(track)])
        print(data_dict)

        return data_dict

    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")
        return {}

filename_list = text_data.get("filename", [])
if not isinstance(filename_list, list):
    filename_list = []


if filename_list:
    tracks = extract_keywords_from_excel(filename_list[0])
else:
    tracks = {}
    

        
@bot.message_handler(commands=['start'])
def main(message):
    
    markup = ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(KeyboardButton("Русский"), KeyboardButton("Тоҷикӣ"))
    bot.send_message(message.chat.id, "Выберите язык!\nЗабонро Интихоб кунед!", reply_markup=markup)
    
@bot.message_handler(func=lambda message: message.text in ["Русский", "Тоҷикӣ"])
def handle_language_choice(message):
    lan = lan_checker(message)
    localized_text = text_data.get(lan, {})  # Достаём локализованный текст
    if lan:
        user_language[str(message.chat.id)] = lan
        text_data["users"]["user_language"] = user_language
        save_data(DATA_FILE, text_data)
        
        greeting = localized_text.get("hello", ["Ошибка"])[0]
        bot.send_message(message.chat.id, f"{greeting}")
        
        
        user = message.from_user.username
        
        markup = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        if user in MODERATORS:
            buttons = ["/start",
                       localized_text['commands']["adminList"],
                       localized_text['commands']["addAdmin"],
                       localized_text['commands']["trackList"],
                       localized_text['commands']["newFile"],
                       localized_text['commands']["tracker"],
                       localized_text['commands']["delAdmin"]]
        elif user in ADMINS:
            buttons = ["/start",
                       localized_text['commands']["trackList"],
                       localized_text['commands']["newFile"],
                       localized_text['commands']["tracker"]]
        else:
            buttons = ["/start", 
                       localized_text['commands']["tracker"],
                       localized_text['commands']["address"],
                       localized_text['commands']["progo"],
                       localized_text['commands']["price"]]
            
            
        for i in range(0, len(buttons), 2):
            markup.row(*[KeyboardButton(btn) for btn in buttons[i:i+2]])
            
        bot.send_message(message.chat.id, localized_text['choose'], reply_markup=markup)
        
    else:
        bot.send_message(message.chat.id, localized_text['lanError'])



@bot.message_handler(func=lambda message: str(message.chat.id) in user_language)
def handle_buttons(message):
    chat_id = str(message.chat.id)
    lan = user_language.get(chat_id, "ru")  # Получаем язык пользователя (по умолчанию "ru")
    localized_text = text_data.get(lan, {}).get("commands", {})
    anss = text_data.get(lan, {}).get("response", {})
    
    if message.text == localized_text.get("addAdmin", "addAdmin"):
        waiting_for_admin[chat_id] = [True, "add"]  # Устанавливаем флаг ожидания имени нового админа
        text_data["users"]["waiting_for_admin"] = waiting_for_admin
        save_data(DATA_FILE, text_data)
        bot.send_message(chat_id, anss.get("addAdmin", anss['addAdmin'][0]))
        return  # Выход из функции, чтобы не обрабатывать дальше
    
    def nameCheck(message):
        arr = list('qwertyuiopasdfghjklzxcvbnm')
        name = list(message)
        for i in name:
            if i not in arr:
                return False
        else:
            return True
    
    # Проверяем, ожидаем ли мы ввод имени нового админа
    if waiting_for_admin.get(chat_id, [False, ""])[1] == "add" and nameCheck(message.text):
        new_admin = message.text.strip().lstrip('@')  # Убираем @, если пользователь его ввел
        if new_admin and new_admin not in ADMINS:
            bot.send_message(chat_id, f"{anss['addAdmin'][1]}{new_admin}{anss['addAdmin'][2]}")
            text_data["users"]["admins"].append(new_admin)  # Добавляем в JSON
            save_data(DATA_FILE, text_data)  # Сохраняем обновленный JSON
        else:
            bot.send_message(chat_id, anss['addAdmin'][-1])
        
        waiting_for_admin.pop(chat_id)  # Убираем флаг ожидания
        return

    if message.text == localized_text.get("delAdmin", "delAdmin"):
        waiting_for_admin[chat_id] = [True, "del"]  # Устанавливаем флаг ожидания имени нового админа
        text_data["users"]["waiting_for_admin"] = waiting_for_admin
        save_data(DATA_FILE, text_data)
        bot.send_message(chat_id, anss.get("delAdmin", anss['delAdmin'][0]))
        return  # Выход из функции, чтобы не обрабатывать дальше
    
    # Проверяем, ожидаем ли мы ввод имени нового админа
    if waiting_for_admin.get(chat_id, [False, ""])[1] == "del":
        new_admin = message.text.strip().lstrip('@')  # Убираем @, если пользователь его ввел
        if new_admin and new_admin in ADMINS:
            bot.send_message(chat_id, f"{anss['delAdmin'][1]}{new_admin}{anss['delAdmin'][2]}")
            text_data["users"]["admins"].remove(new_admin)  # Удаляем из списка
            save_data(DATA_FILE, text_data)  # Сохраняем обновленный JSON
        else:
            bot.send_message(chat_id, anss['delAdmin'][-1])
        
        waiting_for_admin.pop(chat_id)  # Убираем флаг ожидания
        text_data["users"]["waiting_for_admin"] = waiting_for_admin
        save_data(DATA_FILE, text_data)  # Сохраняем обновленный JSON
        return
    
    if message.text == localized_text.get("newFile", "newFile"):
        text_data["users"]["waiting_for_file"] = {chat_id: True}
        save_data(DATA_FILE, text_data)
        info_list = text_data.get(lan, {}).get("file", {}).get("info", [])
        if isinstance(info_list, list) and info_list:
            bot.send_message(chat_id, info_list[0])
        # else:
        #     bot.send_message(chat_id, "Ошибка: отсутствует информация в 'info'.")
    
    if message.text == localized_text.get("tracker", "tracker"):
        waiting_for_admin[chat_id] = [True, "tr"]  # Устанавливаем флаг ожидания имени нового админа
        text_data["users"]["waiting_for_admin"] = waiting_for_admin
        save_data(DATA_FILE, text_data)
        bot.send_message(chat_id, anss.get("tracker", anss['tracker'][0]))
        return
    if waiting_for_admin.get(chat_id, [False, ""])[1] == "tr":
        track = message.text
        value = tracks.get(track)
        if value is not None:
            bot.send_message(chat_id, f"{anss['tracker'][1]}{value}{anss['tracker'][2]}")
        else:
            bot.send_message(chat_id, anss['tracker'][-1])
                
        waiting_for_admin.pop(chat_id)  # Убираем флаг ожидания
        text_data["users"]["waiting_for_admin"] = waiting_for_admin
        save_data(DATA_FILE, text_data)  # Сохраняем обновленный JSON
        return


    response = {
        localized_text.get("adminList", "adminList"): f"{anss['adminList']}\n" + "\n".join(f"@{admin}" for admin in ADMINS),
        localized_text.get("trackList", "trackList"): f"{anss['trackList']}\n" + "\n".join(str(track) for track in tracks),
        localized_text.get("address","address"):anss['address'],
        localized_text.get("progo","progo"):anss["progo"],
        localized_text.get("price","price"):anss["price"]
    }

    if message.text in response:
        bot.send_message(message.chat.id, response[message.text])
    # else:
    #     bot.send_message(message.chat.id, "Неизвестная команда.")
    



@bot.message_handler(content_types=['document'])
def handle_new_file(message):
    chat_id = str(message.chat.id)
    lan = user_language.get(chat_id, "ru")  # Получаем язык пользователя (по умолчанию "ru")
    localized_text = text_data.get(lan, {}).get("file", {})

    # Проверяем, действительно ли пользователь нажал кнопку перед отправкой файла
    if not text_data["users"].get("waiting_for_file", {}).get(chat_id, False):
        bot.send_message(chat_id, "Сначала нажмите кнопку 'Новый файл'.")
        return

    doc: Document = message.document
    if not doc.file_name.endswith('.xlsx'):
        bot.send_message(chat_id, localized_text["errors"]["format"])
        return

    file_info = bot.get_file(doc.file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    file_path = os.path.join(os.getcwd(), doc.file_name)

    with open(file_path, 'wb') as new_file:
        new_file.write(downloaded_file)

    if not validate_xlsx(file_path):
        os.remove(file_path)
        bot.send_message(chat_id, localized_text["errors"]["validation"])
        return

    if filename_list:
        old_file = filename_list.pop()
        if os.path.exists(old_file):
            move_file(old_file, './old_files')

    filename_list.append(doc.file_name)
    text_data["filename"] = filename_list
    save_data(DATA_FILE, text_data)
    
    global tracks
    tracks = extract_keywords_from_excel(os.path.join(os.getcwd(), doc.file_name))

    # Сбрасываем флаг ожидания файла
    text_data["users"]["waiting_for_file"].pop(chat_id, None)
    save_data(DATA_FILE, text_data)

    bot.send_message(chat_id, f"{localized_text['info'][1]} {doc.file_name} {localized_text['info'][2]}")






if __name__ == "__main__":
    while True:
        try:
            bot.polling(none_stop=True, timeout=60, long_polling_timeout=60)
        except Exception as e:
            print(f"Ошибка: {e}")
            time.sleep(5)